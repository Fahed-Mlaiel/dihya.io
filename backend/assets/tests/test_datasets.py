import json, csv, yaml, os
import xml.etree.ElementTree as ET
import toml
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

def test_users_sample():
    with open('users_sample.json') as f:
        data = json.load(f)
    assert all('email' in u and '@' in u['email'] for u in data)
    assert all(u['lang'] in ['fr','en','ar','kab'] for u in data)
    assert all(u['username'].startswith('user') for u in data)

def test_transactions_sample():
    with open('transactions_sample.csv') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert all(float(r['amount']) >= 0 for r in rows)
    assert all(r['currency'] in ['EUR','USD','DZD'] for r in rows)
    assert all(r['status'] in ['completed','failed','pending'] for r in rows)

def test_audit_events_sample():
    with open('audit_events_sample.yaml') as f:
        events = yaml.safe_load(f)
    assert all('event' in e for e in events)
    assert all(e['status'] in ['success','denied','error'] for e in events)

def test_roles_sample():
    with open('roles_sample.json') as f:
        roles = json.load(f)
    ids = [r['id'] for r in roles]
    assert len(ids) == len(set(ids)), 'IDs non uniques dans roles_sample.json'
    assert all(r['lang'] in ['fr','en','ar','kab'] for r in roles)
    assert all(r['role'] in ['admin','user','auditor','moderator'] for r in roles)

def test_permissions_sample():
    with open('permissions_sample.yaml') as f:
        perms = yaml.safe_load(f)
    ids = [p['id'] for p in perms]
    assert len(ids) == len(set(ids)), 'IDs non uniques dans permissions_sample.yaml'
    assert all(p['role'] in ['admin','user','auditor','moderator'] for p in perms)
    assert all(p['permission'] in ['all','read','audit','moderate'] for p in perms)

def test_projects_sample():
    with open('projects_sample.json') as f:
        data = json.load(f)
    assert all('name' in p for p in data)
    assert all(p['status'] in ['active','archived','pending'] for p in data)
    assert all(p['lang'] in ['fr','en','ar','kab'] for p in data)

def test_consent_events_sample():
    with open('consent_events_sample.yaml') as f:
        events = yaml.safe_load(f)
    assert all(e['status'] in ['given','refused','withdrawn'] for e in events)
    assert all(e['consent_type'] in ['export','marketing','profiling'] for e in events)

def test_notifications_sample():
    with open('notifications_sample.csv') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert all(r['type'] in ['incident','role','rgpd'] for r in rows)
    assert all(r['lang'] in ['fr','en','ar','kab'] for r in rows)
    assert all(r['read'] in ['true','false'] for r in rows)

def test_rgpd_anonymisation():
    # Vérifie qu'aucun nom réel ou email réel n'est présent
    for fname in os.listdir('.'):
        if fname.endswith(('.json','.csv','.yaml')):
            with open(fname, encoding='utf-8') as f:
                content = f.read().lower()
                assert not any(x in content for x in ['dupont','durand','smith','@gmail.com','@yahoo.com','adresse','address','téléphone','phone','numéro']), f"Donnée personnelle détectée dans {fname}"

def test_logs_sample():
    with open('logs_sample.json') as f:
        logs = json.load(f)
    assert all('timestamp' in l and 'level' in l for l in logs)
    assert all(l['level'] in ['INFO','ERROR','WARNING','DEBUG'] for l in logs)

def test_roles_sample():
    with open('roles_sample.csv') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert all(r['role_name'] in ['admin','user'] for r in rows)
    assert all(r['lang'] in ['fr','en','ar','kab'] for r in rows)

def test_policies_sample():
    with open('policies_sample.yaml') as f:
        policies = yaml.safe_load(f)
    assert all('name' in p and 'active' in p for p in policies)

def test_metrics_sample():
    with open('metrics_sample.json') as f:
        metrics = json.load(f)
    assert all('date' in m and 'active_users' in m for m in metrics)

def test_exports_sample():
    with open('exports_sample.csv') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert all(r['type'] in ['rgpd','accessibility'] for r in rows)
    assert all(r['status'] in ['success','failed'] for r in rows)

def test_transactions_sample_xml():
    tree = ET.parse('transactions_sample.xml')
    root = tree.getroot()
    assert root.tag == 'transactions'
    for t in root.findall('transaction'):
        assert t.attrib['status'] in ['completed','failed','pending']

def test_consent_events_sample_toml():
    data = toml.load('consent_events_sample.toml')
    assert 'consent_events' in data
    for e in data['consent_events']:
        assert e['status'] in ['given','refused','withdrawn']

def test_notifications_sample_md():
    with open('notifications_sample.md') as f:
        content = f.read()
    assert '| id |' in content and '| user_id |' in content

def test_users_sample_xlsx():
    if not openpyxl:
        print('openpyxl non installé, test ignoré')
        return
    wb = openpyxl.load_workbook('users_sample.xlsx')
    ws = wb.active
    assert ws.max_row > 1
    assert ws['A1'].value.lower() == 'id'

def test_audit_events_sample_pdf():
    if not PyPDF2:
        print('PyPDF2 non installé, test ignoré')
        return
    with open('audit_events_sample.pdf', 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        assert len(reader.pages) > 0

if __name__ == '__main__':
    test_users_sample()
    test_transactions_sample()
    test_audit_events_sample()
    test_roles_sample()
    test_permissions_sample()
    test_projects_sample()
    test_consent_events_sample()
    test_notifications_sample()
    test_rgpd_anonymisation()
    test_logs_sample()
    test_roles_sample()
    test_policies_sample()
    test_metrics_sample()
    test_exports_sample()
    test_transactions_sample_xml()
    test_consent_events_sample_toml()
    test_notifications_sample_md()
    test_users_sample_xlsx()
    test_audit_events_sample_pdf()
    print('All dataset tests passed.')
